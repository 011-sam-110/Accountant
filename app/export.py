"""Build the accountant handover pack: PDF summary + Excel + CSV + receipts → zip.

Called from app.routes.export. Returns bytes (a .zip). Works even for an
empty year (zeros, empty entries sheet). Receipt failures are silently skipped
so a missing image never breaks the whole pack.
"""
from __future__ import annotations

import calendar
import csv
import io
import logging
import re
import zipfile
from datetime import date

from app.util import format_pounds

log = logging.getLogger("mtd.export")


# ----------------------------------------------------------------- helpers ----

def safe_filename(s: str) -> str:
    """Strip characters that are unsafe in zip entry names / OS filenames."""
    s = (s or "unknown").strip()
    s = re.sub(r"[^\w\-. ]", "_", s)
    s = re.sub(r"\s+", "_", s)
    return s[:60] or "unknown"


def _nicedate(d: date) -> str:
    return f"{d.day} {calendar.month_abbr[d.month]} {d.year}"


# ---------------------------------------------------------------- PDF -------

def _build_pdf(repo, user, tax_year) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=f"Accountant pack — {tax_year.label}",
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    heading1 = ParagraphStyle(
        "H1",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=6,
        textColor=colors.HexColor("#1a7a3c"),
    )
    heading2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=14,
        spaceAfter=4,
        textColor=colors.HexColor("#1a7a3c"),
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=normal,
        fontSize=9,
        textColor=colors.HexColor("#666666"),
    )
    note_style = ParagraphStyle(
        "Note",
        parent=normal,
        fontSize=9,
        textColor=colors.HexColor("#555555"),
        leftIndent=0,
        spaceAfter=4,
    )

    # ---- table style helpers ----
    def _tbl(data, col_widths, header_rows=1):
        t = Table(data, colWidths=col_widths)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f5ec")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1a7a3c")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f9f9f9")]),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
            ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ]
        t.setStyle(TableStyle(style))
        return t

    story = []

    # ---- title block ----
    story.append(Paragraph(f"Accountant pack — {tax_year.label}", heading1))
    story.append(Paragraph(
        f"<b>Prepared for:</b> {user.display_name or user.email} "
        f"({user.email})", normal))
    story.append(Paragraph(
        f"<b>Generated:</b> {_nicedate(date.today())}", normal))
    story.append(Paragraph(
        f"<b>Period:</b> {_nicedate(tax_year.start_date)} "
        f"to {_nicedate(tax_year.end_date)}", normal))

    if tax_year.status == "closed" and tax_year.snapshot:
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "Figures frozen on closing.", meta_style))

    story.append(Spacer(1, 14))

    # ---- totals block ----
    story.append(Paragraph("Summary", heading2))

    gross = repo.gross_income_minor(tax_year.id)
    expense = repo.expense_total_minor(tax_year.id)
    net = gross - expense

    totals_data = [
        ["", "Amount"],
        ["Total income", format_pounds(gross)],
        ["Total expenses", format_pounds(expense)],
        ["Net profit", format_pounds(net)],
    ]
    story.append(_tbl(totals_data, [11 * cm, 4 * cm]))
    story.append(Spacer(1, 14))

    # ---- SA103 boxes table ----
    boxes = repo.totals_by_box(tax_year.id)
    if boxes:
        story.append(Paragraph("SA103 box totals", heading2))

        # Split income vs expense
        income_rows = [(b["box_number"], b["box_label"], format_pounds(b["total_minor"]))
                       for b in boxes.values() if b["kind"] == "income"]
        expense_rows = [(b["box_number"], b["box_label"], format_pounds(b["total_minor"]))
                        for b in boxes.values() if b["kind"] == "expense"]

        if income_rows:
            story.append(Paragraph("Income", ParagraphStyle(
                "sub", parent=normal, fontSize=10, textColor=colors.HexColor("#1a7a3c"),
                spaceBefore=6, spaceAfter=2)))
            data = [["Box", "Description", "Amount"]] + list(income_rows)
            story.append(_tbl(data, [1.5 * cm, 9.5 * cm, 4 * cm]))
            story.append(Spacer(1, 8))

        if expense_rows:
            story.append(Paragraph("Expenses", ParagraphStyle(
                "sub2", parent=normal, fontSize=10, textColor=colors.HexColor("#c0392b"),
                spaceBefore=6, spaceAfter=2)))
            data = [["Box", "Description", "Amount"]] + list(expense_rows)
            story.append(_tbl(data, [1.5 * cm, 9.5 * cm, 4 * cm]))
            story.append(Spacer(1, 8))

    # ---- category breakdown ----
    cat_totals = repo.category_totals(tax_year.id)
    if cat_totals:
        story.append(Paragraph("Category breakdown", heading2))
        data = [["Category", "Maps to box", "Amount"]]
        for row in cat_totals:
            data.append([
                row["label"],
                f"{row['box_number']} — {row['box_label']}" if row["box_label"] else row["box_number"],
                format_pounds(row["total_minor"]),
            ])
        story.append(_tbl(data, [6 * cm, 7 * cm, 4 * cm]))
        story.append(Spacer(1, 14))

    # ---- footer note ----
    story.append(Spacer(1, 20))
    story.append(Paragraph(
        "Figures are SA103-aligned. Please confirm the exact box numbers for the "
        "current year's form.",
        note_style))
    story.append(Paragraph(
        "Generated by EasyBooks MTD.", meta_style))

    doc.build(story)
    return buf.getvalue()


# --------------------------------------------------------------- Excel -------

def _build_xlsx(repo, tax_year) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ---- Entries sheet ----
    ws = wb.active
    ws.title = "Entries"

    header = ["Date", "Type", "Category", "SA103 Box", "Vendor/Student",
              "Amount (£)", "Paid?"]
    ws.append(header)

    # Style header row
    header_fill = PatternFill("solid", fgColor="e8f5ec")
    header_font = Font(bold=True, color="1a7a3c")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    entries = repo.list_entries(tax_year_id=tax_year.id)
    for e in entries:
        category_label = e.category.label if e.category else ""
        # resolve SA103 box number
        box_number = ""
        if e.category and hasattr(e.category, "box") and e.category.box:
            box_number = e.category.box.box_number
        # vendor or student name
        vendor_or_student = (
            e.student.name if e.student else (e.vendor or "")
        )
        paid_str = ""
        if e.entry_type == "income":
            paid_str = "Yes" if e.is_paid else "No"
        ws.append([
            e.entry_date.isoformat(),
            e.entry_type,
            category_label,
            box_number,
            vendor_or_student,
            round(e.amount_minor / 100, 2),
            paid_str,
        ])

    # Widen columns for readability
    col_widths = [12, 10, 24, 12, 28, 12, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Summary sheet ----
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Box", "Label", "Kind", "Amount (£)"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    boxes = repo.totals_by_box(tax_year.id)
    total_income = 0
    total_expense = 0
    for b in boxes.values():
        ws2.append([
            b["box_number"],
            b["box_label"],
            b["kind"],
            round(b["total_minor"] / 100, 2),
        ])
        if b["kind"] == "income":
            total_income += b["total_minor"]
        else:
            total_expense += b["total_minor"]

    ws2.append([])
    ws2.append(["", "Total income", "", round(total_income / 100, 2)])
    ws2.append(["", "Total expenses", "", round(total_expense / 100, 2)])
    ws2.append(["", "Net", "", round((total_income - total_expense) / 100, 2)])

    for i, w in enumerate([10, 36, 10, 14], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------- CSV ------

def _build_csv(repo, tax_year) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Type", "Category", "SA103 Box",
                     "Vendor/Student", "Amount (£)", "Paid?"])

    entries = repo.list_entries(tax_year_id=tax_year.id)
    for e in entries:
        category_label = e.category.label if e.category else ""
        box_number = ""
        if e.category and hasattr(e.category, "box") and e.category.box:
            box_number = e.category.box.box_number
        vendor_or_student = e.student.name if e.student else (e.vendor or "")
        paid_str = ""
        if e.entry_type == "income":
            paid_str = "Yes" if e.is_paid else "No"
        writer.writerow([
            e.entry_date.isoformat(),
            e.entry_type,
            category_label,
            box_number,
            vendor_or_student,
            f"{e.amount_minor / 100:.2f}",
            paid_str,
        ])
    return buf.getvalue().encode("utf-8")


# -------------------------------------------------------------- README ------

_README = """\
EasyBooks — Accountant Handover Pack
======================================

This pack contains everything your accountant needs to complete your
Self Assessment (SA103 supplementary pages) for the tax year shown.

Files in this pack:

  summary.pdf
      A human-readable summary of your income, expenses, and net profit,
      plus a breakdown by SA103 tax box and by category. Use this as the
      cover sheet when handing the pack over.

  data.xlsx
      A spreadsheet with two sheets:
        • Entries — every transaction for the year, with date, type,
          category, SA103 box, vendor or student name, amount, and (for
          income entries) whether it has been marked as paid.
        • Summary — the SA103 box totals and an overall income/expense/net.

  data.csv
      The same Entries data as a plain CSV, useful if your accountant prefers
      to import it directly into their own software.

  receipts/ (folder)
      Photos of receipts, named by date and vendor/category. These are the
      original images captured in the app. A missing receipt for any
      individual entry is silently omitted — please check the app if any
      are missing.

Notes:
  • All figures are in pounds sterling (GBP).
  • Figures are SA103-aligned. Please confirm the exact box numbers for the
    current year's form, as HMRC can renumber boxes between tax years.
  • If the tax year was closed in the app, figures are frozen and cannot
    change after the closing date.
"""


# ============================================================= main entry ====

def build_pack(repo, user, tax_year) -> bytes:
    """Build and return the complete handover pack as zip bytes."""
    from app.storage import get_storage

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:

        # --- summary.pdf ---
        try:
            pdf_bytes = _build_pdf(repo, user, tax_year)
            zf.writestr("summary.pdf", pdf_bytes)
        except Exception:
            log.exception("export: PDF build failed")
            zf.writestr("summary.pdf", b"(PDF generation failed)")

        # --- data.xlsx ---
        try:
            xlsx_bytes = _build_xlsx(repo, tax_year)
            zf.writestr("data.xlsx", xlsx_bytes)
        except Exception:
            log.exception("export: XLSX build failed")

        # --- data.csv ---
        try:
            csv_bytes = _build_csv(repo, tax_year)
            zf.writestr("data.csv", csv_bytes)
        except Exception:
            log.exception("export: CSV build failed")

        # --- receipts/ ---
        storage = get_storage()
        entries = repo.list_entries(tax_year_id=tax_year.id)
        for e in entries:
            try:
                receipt = repo.receipt_for_entry(e.id)
                if receipt is None:
                    continue
                image_bytes = storage.get_bytes(receipt.storage_key)
                # determine extension from content_type
                ext = "jpg"
                ct = (receipt.content_type or "image/jpeg").lower()
                if "png" in ct:
                    ext = "png"
                elif "gif" in ct:
                    ext = "gif"
                vendor_or_cat = (
                    e.vendor
                    or (e.category.label if e.category else "")
                    or "entry"
                )
                filename = (
                    f"receipts/{e.entry_date:%Y-%m-%d}"
                    f"_{safe_filename(vendor_or_cat)}"
                    f"_{e.id[:8]}.{ext}"
                )
                zf.writestr(filename, image_bytes)
            except Exception:
                # Missing or unreadable image — skip gracefully
                continue

        # --- README.txt ---
        zf.writestr("README.txt", _README.encode("utf-8"))

    return buf.getvalue()
